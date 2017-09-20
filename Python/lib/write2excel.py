"""
This module implements a class to create Excel Workbooks.
"""

import logging
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side


class Write2Excel:
    """
    This class consolidates the functions to write the data to excel workbook a nice format.
    """

    def __init__(self):
        """
        Create the workbook object and set the initial sheet.
        """
        self.wb = Workbook()
        self.current_sheet = self.wb.active
        self.rowcnt = 0

    def init_sheet(self, title):
        """
        Initialize the worksheet.

        :param title: mandatory title for the worksheet

        :return: worksheet object
        """
        if self.current_sheet.title == 'Sheet':
            # This is the initial sheet, re-assign title
            self.current_sheet.title = title
        else:
            # There are sheets already, finalize the sheet and create a new one
            finalize_sheet(self.current_sheet, self.rowcnt)
            self.current_sheet = self.wb.create_sheet(title=title)
        initialize_sheet(self.current_sheet)
        self.rowcnt = 2
        return self.current_sheet

    def init_sheet_cons(self, title):
        """
        Initialize the worksheet for Consolidation.

        :param title: mandatory title for the worksheet

        :return: worksheet object
        """
        self.current_sheet.title = title
        initialize_sheet_report_combis(self.current_sheet)
        self.rowcnt = 2
        return self.current_sheet

    def init_sheet_cons_stap(self, title):
        """
        Initialize the worksheet for Consolidation.

        :param title: mandatory title for the worksheet

        :return: worksheet object
        """
        self.current_sheet.title = title
        initialize_sheet_report_combis_stap(self.current_sheet)
        self.rowcnt = 2
        return self.current_sheet

    def close_workbook(self, filename):
        """
        This method will finalize the current sheet and close the file.
        :param filename:
        :return:
        """
        finalize_sheet(self.current_sheet, self.rowcnt)
        self.wb.save(filename=filename)
        return

    def close_workbook_cons(self, filename):
        """
        This method will finalize the current sheet and close the file.
        :param filename:
        :return:
        """
        finalize_sheet_cons(self.current_sheet, self.rowcnt)
        self.wb.save(filename=filename)
        return

    def close_workbook_cons_stap(self, filename):
        """
        This method will finalize the current sheet and close the file.
        :param filename:
        :return:
        """
        finalize_sheet_cons_stap(self.current_sheet, self.rowcnt)
        self.wb.save(filename=filename)
        return

    def write_line(self, content):

        if 'Laatste Aanleg' in content['arfase']:
            bg_color = 'FFC000'     # Orange
        elif ('Samenstellen' in content['arfase']) or ('Dossier' in content['arstap']):
            # Fase: Eerste Aanleg - Stap: Samenstellen of Indienen Dossier.
            bg_color = 'FFFF00'     # Yellow
        elif 'Uitvoeren' in content['arfase']:
            # Fase: Uitvoeren
            bg_color = '00B0F0'  # Blue
        elif 'Eerste Aanleg' in content['arfase']:
            bg_color = '92D050'     # Green
        else:
            bg_color = 'FF0000'  # Red

        fill_cell = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        self.rowcnt += 1
        self.current_sheet[rc2a1(self.rowcnt, 1)] = content['decreet']
        self.current_sheet[rc2a1(self.rowcnt, 2)] = content['besluit']
        self.current_sheet[rc2a1(self.rowcnt, 3)] = content['arstap']
        self.current_sheet[rc2a1(self.rowcnt, 4)] = content['ardoc']
        self.current_sheet[rc2a1(self.rowcnt, 5)] = content['updoc_code']
        self.current_sheet[rc2a1(self.rowcnt, 6)] = content['updoc']
        self.current_sheet[rc2a1(self.rowcnt, 7)] = content['upgeb_code']
        self.current_sheet[rc2a1(self.rowcnt, 8)] = content['gebeurtenis']
        for cnt in range(1, 9):
            self.current_sheet[rc2a1(self.rowcnt, cnt)].fill = fill_cell
        return

    def write_line4omer(self, content):
        if 'laatste aanleg' in content['upfase']:
            bg_color = 'FFC000'  # Orange
        elif 'Samenstellen' in content['upfase']:
            # Fase: Eerste Aanleg - Stap: Samenstellen of Indienen Dossier.
            bg_color = 'FFFF00'  # Yellow
        elif 'Uitvoeren' in content['upfase']:
            # Fase: Uitvoeren
            bg_color = '00B0F0'  # Blue
        elif 'eerste aanleg' in content['upfase']:
            bg_color = '92D050'  # Green
        else:
            logging.error("Unexpected UpFase {uf}, no colouring...".format(uf=content['upfase']))
            bg_color = 'FF0000'  # Green

        fill_cell = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        self.rowcnt += 1
        self.current_sheet[rc2a1(self.rowcnt, 1)] = content['decreet']
        self.current_sheet[rc2a1(self.rowcnt, 2)] = content['besluit']
        self.current_sheet[rc2a1(self.rowcnt, 3)] = content['arstap']
        self.current_sheet[rc2a1(self.rowcnt, 4)] = content['ardoc']
        self.current_sheet[rc2a1(self.rowcnt, 5)] = content['updoc_code']
        self.current_sheet[rc2a1(self.rowcnt, 6)] = content['updoc']
        self.current_sheet[rc2a1(self.rowcnt, 7)] = content['upgeb_code']
        self.current_sheet[rc2a1(self.rowcnt, 8)] = content['gebeurtenis']
        for cnt in range(1, 9):
            self.current_sheet[rc2a1(self.rowcnt, cnt)].fill = fill_cell
        return

    def write_line_report_combis(self, content):
        """
        This function will write a line for each unique combination per source.
        :param self:
        :param content:
        :return:
        """
        if 'HanVwpFun' in content['bron']:
            bg_color = '00B0F0'  # Blue
        elif 'HanVwp' in content['bron']:
            bg_color = '92D050'  # Green
        elif 'Vast' in content['bron']:
            bg_color = 'FFD1DC'
        elif 'Milieu' in content['bron']:
            bg_color = 'FFFF00'  # Yellow
        else:
            bg_color = 'FFC000'  # Orange

        fill_cell = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        self.rowcnt += 1
        self.current_sheet[rc2a1(self.rowcnt, 1)] = content['bron']
        self.current_sheet[rc2a1(self.rowcnt, 2)] = content['dt_code']
        self.current_sheet[rc2a1(self.rowcnt, 3)] = content['dt_naam']
        self.current_sheet[rc2a1(self.rowcnt, 4)] = content['f_code']
        self.current_sheet[rc2a1(self.rowcnt, 5)] = content['f_naam']
        self.current_sheet[rc2a1(self.rowcnt, 6)] = content['g_code']
        self.current_sheet[rc2a1(self.rowcnt, 7)] = content['g_naam']
        self.current_sheet[rc2a1(self.rowcnt, 8)] = content['d_type']
        self.current_sheet[rc2a1(self.rowcnt, 9)] = content['d_code']
        self.current_sheet[rc2a1(self.rowcnt, 10)] = content['d_naam']
        for cnt in range(1, 11):
            self.current_sheet[rc2a1(self.rowcnt, cnt)].fill = fill_cell
        return

    def write_line_report_combis_stap(self, content):
        """
        This function will write a line for each unique combination per source.
        :param self:
        :param content:
        :return:
        """
        if 'HanVwpFun' in content['bron']:
            bg_color = '00B0F0'  # Blue
        elif 'HanVwp' in content['bron']:
            bg_color = '92D050'  # Green
        elif 'Vast' in content['bron']:
            bg_color = 'FFD1DC'
        elif 'Milieu' in content['bron']:
            bg_color = 'FFFF00'  # Yellow
        else:
            bg_color = 'FFC000'  # Orange

        fill_cell = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        self.rowcnt += 1
        self.current_sheet[rc2a1(self.rowcnt, 1)] = content['bron']
        self.current_sheet[rc2a1(self.rowcnt, 2)] = content['dt_code']
        self.current_sheet[rc2a1(self.rowcnt, 3)] = content['dt_naam']
        self.current_sheet[rc2a1(self.rowcnt, 4)] = content['f_code']
        self.current_sheet[rc2a1(self.rowcnt, 5)] = content['f_naam']
        self.current_sheet[rc2a1(self.rowcnt, 6)] = content['g_code']
        self.current_sheet[rc2a1(self.rowcnt, 7)] = content['g_naam']
        self.current_sheet[rc2a1(self.rowcnt, 8)] = content['s_naam']
        self.current_sheet[rc2a1(self.rowcnt, 9)] = content['d_type']
        self.current_sheet[rc2a1(self.rowcnt, 10)] = content['d_code']
        self.current_sheet[rc2a1(self.rowcnt, 11)] = content['d_naam']
        for cnt in range(1, 12):
            self.current_sheet[rc2a1(self.rowcnt, cnt)].fill = fill_cell
        return


def initialize_sheet(ws):
    # Set column width
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 32
    ws.column_dimensions[get_column_letter(4)].width = 72
    ws.column_dimensions[get_column_letter(5)].width = 32
    ws.column_dimensions[get_column_letter(6)].width = 72
    ws.column_dimensions[get_column_letter(7)].width = 40
    ws.column_dimensions[get_column_letter(8)].width = 100

    # Set row height
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # Create Title Row 1
    bg_color = 'BFBFBF'
    fill_r1 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:D1')
    ws.merge_cells('E1:H1')
    ws['A1'] = 'Wetgeving Artikels'
    ws['C1'] = 'Register'
    ws['E1'] = 'Omgevingsloket'
    ws['A1'].fill = fill_r1
    ws['C1'].fill = fill_r1
    ws['E1'].fill = fill_r1
    set_bold(ws['A1'])
    set_bold(ws['C1'])
    set_bold(ws['E1'])
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    ws['E1'].alignment = Alignment(horizontal='center')

    # Create Title Row 2
    bg_color = 'D9D9D9'
    fill_r2 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    title_row = ['decreet', 'besluit', 'stap', 'document - archief', 'document code',
                 'document - loket', 'gebeurtenis code', 'gebeurtenis']
    row = 2
    for pos in range(len(title_row)):
        col = pos + 1
        ws[rc2a1(row, col)] = title_row[pos]
        ws[rc2a1(row, col)].fill = fill_r2
        set_bold(ws[rc2a1(row, col)])


def initialize_sheet_report_combis(ws):
    # Set column width
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 28
    ws.column_dimensions[get_column_letter(3)].width = 32
    ws.column_dimensions[get_column_letter(4)].width = 28
    ws.column_dimensions[get_column_letter(5)].width = 32
    ws.column_dimensions[get_column_letter(6)].width = 28
    ws.column_dimensions[get_column_letter(7)].width = 32
    ws.column_dimensions[get_column_letter(8)].width = 11
    ws.column_dimensions[get_column_letter(9)].width = 28
    ws.column_dimensions[get_column_letter(10)].width = 64

    # Set row height
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # Create Title Row 1
    bg_color = 'BFBFBF'
    fill_r1 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')
    ws.merge_cells('H1:J1')
    ws['A1'] = 'Bron'
    ws['B1'] = 'Dossier Type'
    ws['D1'] = 'Fase'
    ws['F1'] = 'Gebeurtenis'
    ws['H1'] = 'Document'
    ws['A1'].fill = fill_r1
    ws['B1'].fill = fill_r1
    ws['D1'].fill = fill_r1
    ws['F1'].fill = fill_r1
    ws['H1'].fill = fill_r1
    set_bold(ws['A1'])
    set_bold(ws['B1'])
    set_bold(ws['D1'])
    set_bold(ws['F1'])
    set_bold(ws['H1'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='bottom')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['F1'].alignment = Alignment(horizontal='center')
    ws['H1'].alignment = Alignment(horizontal='center')

    # Create Title Row 2
    bg_color = 'D9D9D9'
    fill_r2 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    title_row = ['Code', 'Naam', 'Code', 'Naam', 'Code', 'Naam', 'Type', 'Code', 'Naam']
    row = 2
    for pos in range(len(title_row)):
        col = pos + 2
        ws[rc2a1(row, col)] = title_row[pos]
        ws[rc2a1(row, col)].fill = fill_r2
        set_bold(ws[rc2a1(row, col)])


def initialize_sheet_report_combis_stap(ws):
    # Set column width
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 28
    ws.column_dimensions[get_column_letter(3)].width = 32
    ws.column_dimensions[get_column_letter(4)].width = 28
    ws.column_dimensions[get_column_letter(5)].width = 32
    ws.column_dimensions[get_column_letter(6)].width = 28
    ws.column_dimensions[get_column_letter(7)].width = 32
    ws.column_dimensions[get_column_letter(8)].width = 32
    ws.column_dimensions[get_column_letter(9)].width = 11
    ws.column_dimensions[get_column_letter(10)].width = 28
    ws.column_dimensions[get_column_letter(11)].width = 64

    # Set row height
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # Create Title Row 1
    bg_color = 'BFBFBF'
    fill_r1 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')
    ws.merge_cells('I1:K1')
    ws['A1'] = 'Bron'
    ws['B1'] = 'Dossier Type'
    ws['D1'] = 'Fase'
    ws['F1'] = 'Gebeurtenis'
    ws['H1'] = 'Stap'
    ws['I1'] = 'Document'
    ws['A1'].fill = fill_r1
    ws['B1'].fill = fill_r1
    ws['D1'].fill = fill_r1
    ws['F1'].fill = fill_r1
    ws['H1'].fill = fill_r1
    ws['I1'].fill = fill_r1
    set_bold(ws['A1'])
    set_bold(ws['B1'])
    set_bold(ws['D1'])
    set_bold(ws['F1'])
    set_bold(ws['H1'])
    set_bold(ws['I1'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='bottom')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['F1'].alignment = Alignment(horizontal='center')
    ws['H1'].alignment = Alignment(horizontal='center')
    ws['I1'].alignment = Alignment(horizontal='center')

    # Create Title Row 2
    bg_color = 'D9D9D9'
    fill_r2 = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    title_row = ['Code', 'Naam', 'Code', 'Naam', 'Code', 'Naam', 'Stap', 'Type', 'Code', 'Naam']
    row = 2
    for pos in range(len(title_row)):
        col = pos + 2
        ws[rc2a1(row, col)] = title_row[pos]
        ws[rc2a1(row, col)].fill = fill_r2
        set_bold(ws[rc2a1(row, col)])


def finalize_sheet(ws, nr_rows):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws["A1:H{rc}".format(rc=nr_rows)]:
        for cell in row:
            cell.border = thin_border
    ws.auto_filter.ref = "A2:H{rc}".format(rc=nr_rows)
    ws.freeze_panes = ws['A3']


def finalize_sheet_cons(ws, nr_rows):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws["A1:J{rc}".format(rc=nr_rows)]:
        for cell in row:
            cell.border = thin_border
    ws.auto_filter.ref = "A2:J{rc}".format(rc=nr_rows)
    ws.freeze_panes = ws['A3']


def finalize_sheet_cons_stap(ws, nr_rows):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws["A1:K{rc}".format(rc=nr_rows)]:
        for cell in row:
            cell.border = thin_border
    ws.auto_filter.ref = "A2:K{rc}".format(rc=nr_rows)
    ws.freeze_panes = ws['A3']


def rc2a1(row=None, col=None):
    """
    This function converts a (row, column) pair (R1C1 notation) to the A1 string notation for excel. The column number
    is converted to the character(s), the row number is appended to the column string.

    :param row: Row number (1 .. )

    :param col: Column number (1 .. )

    :return: A1 Notation (column-row) (e.g. 'BF19745')
    """
    return "{0}{1}".format(get_column_letter(col), row)


def set_bold(cell):
    """
    This function will set the font of the cell to bold. In openpyxl cell styles are shared between objects and once
    assigned they cannot be changed anymore. This stops unwanted side-effects such as changing lots of cells instead of
    one single cell.

    :param cell: Font for this cell needs to be set to bold.
    :return: Nothing, but the font of the cell has been changed to bold in the mean time.
    """
    this_font = cell.font
    that_font = copy(this_font)
    that_font.bold = True
    cell.font = that_font
    return
