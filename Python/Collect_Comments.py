"""
This script will collect comments that require attention. This means a search is done on "???"
The comment file is then written to an excel file.
"""
import os
from lib import my_env
from lib import neostore
from openpyxl import Workbook


if __name__ == "__main__":
    cfg = my_env.init_env("convert_protege", __file__)
    ns = neostore.NeoStore(cfg)
    dest_filedir = cfg['Main']['reportdir']
    dest_filename = os.path.join(dest_filedir, "comments.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Commentaar"
    res = ns.get_commentaar()
    row = 1
    col = 1
    for line in res:
        col = 1
        for k, v in line.items():
            if  k == "commentaar":
                ws.cell(row=row, column=col).style.alignment.wrap_text = True
            if k == "labels":
                d = ws.cell(row=row, column=col, value='"{val}"'.format(val=v))
            else:
                d = ws.cell(row=row, column=col, value=v)
            if  k == "commentaar":
                ws.cell(row=row, column=col).style.alignment.wrap_text = True
            col += 1
        row += 1
    wb.save(filename=dest_filename)
